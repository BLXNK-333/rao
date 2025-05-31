# Это временный модуль , потом будет удален

from typing import List
import calendar

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_russian_period_text(month: int, year: int) -> str:
    month_names = {
        1: "января", 2: "февраля", 3: "марта", 4: "апреля",
        5: "мая", 6: "июня", 7: "июля", 8: "августа",
        9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
    }

    from_day = 1
    to_day = calendar.monthrange(year, month)[1]
    month_name = month_names[month]

    return f"{from_day} {month_name} {year} по {to_day} {month_name} {year}"


def generate_month_report(month: int, year: int, data: List[List[str]], filename: str):
    # Инициализация книги и листа
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Шрифты
    header_font = Font(name='Times New Roman', size=14, bold=True)
    subheader_font = Font(name='Times New Roman', size=12, bold=True)
    body_font = Font(name='Times New Roman', size=12)

    # Границы
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Ширина колонок
    col_widths = [48.71, 47.14, 44.14, 6.71, 44.0, 65.71]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Дата диапазон
    period_text = get_russian_period_text(month, year)

    # Шапка
    header_text = (
        f"Отчет  об использованных фонограммах за период {period_text} в эфире радиоканала "
        "« Радио России Астрахань » - город федерального, регионального, республиканского "
        "либо краевого значения осуществляющего трансляцию"
    )
    ws.merge_cells("A1:F1")
    ws["A1"] = header_text
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 54  # увеличено, чтобы вместить многострочный текст

    # Заголовки таблицы
    table_headers = [
        "Название фонограммы",
        "Автор музыки",
        "Автор слов",
        "Кол - во сообщений в эфир",
        "Исполнитель (ФИО исполнителя или названия коллектива)",
        "Изготовитель фонограммы"
    ]
    for col, title in enumerate(table_headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 61.5  # под заголовки с переносом

    # Данные
    for row_idx, row_data in enumerate(data, start=3):
        ws.row_dimensions[row_idx].height = 29
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

    # Пустая строка отступа перед подвалом
    empty_row = len(data) + 3
    ws.row_dimensions[empty_row].height = 22

    # Подвал
    footer_lines = [
        "* ВГТРК ГТРК «Культура» (правопреемник Государственного дома радиовещания и "
        "звукозаписи ГДРЗ)",
        "* Self - Release –  творческая продукция исполнителя, доступная для продаж "
        "или распространения"
    ]
    for i, line in enumerate(footer_lines, start=empty_row + 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = body_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.row_dimensions[i].height = 22

    # Отступ после подвала — 2 строки по 22 pt
    post_footer_row1 = empty_row + len(footer_lines) + 1
    post_footer_row2 = post_footer_row1 + 1
    ws.row_dimensions[post_footer_row1].height = 22
    ws.row_dimensions[post_footer_row2].height = 22

    # Подписи
    signature_row = post_footer_row2 + 1
    ws.row_dimensions[signature_row].height = 22

    ws.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row,
                   end_column=3)  # A:C
    ws.merge_cells(start_row=signature_row, start_column=4, end_row=signature_row,
                   end_column=6)  # D:F

    ws.cell(row=signature_row, column=1,
            value="М.П.    _________________________   ").font = Font(
        name='Times New Roman', size=14, bold=True)
    ws.cell(row=signature_row, column=1).alignment = Alignment(horizontal="center",
                                                               vertical="bottom")

    ws.cell(row=signature_row, column=4, value="Директор ГТРК" + " " * 42).font = Font(
        name='Times New Roman', size=14, bold=True, underline="single")
    ws.cell(row=signature_row, column=4).alignment = Alignment(horizontal="center",
                                                               vertical="bottom")

    # Подписи: поясняющий текст
    explain_row = signature_row + 1
    ws.merge_cells(start_row=explain_row, start_column=1, end_row=explain_row,
                   end_column=3)  # A:C
    ws.merge_cells(start_row=explain_row, start_column=4, end_row=explain_row,
                   end_column=6)  # D:F

    ws.cell(row=explain_row, column=1, value="(подпись)").font = Font(
        name='Times New Roman', size=8)
    ws.cell(row=explain_row, column=1).alignment = Alignment(horizontal="center",
                                                             vertical="top")

    ws.cell(row=explain_row, column=4,
            value="(должность,  ФИО руководителя)").font = Font(
        name='Times New Roman', size=8)
    ws.cell(row=explain_row, column=4).alignment = Alignment(horizontal="center",
                                                             vertical="top")

    # Сохранение
    wb.save(filename)


if __name__ == '__main__':
    raw_data = """Черемшина;Михайлюк В.П.;Юрійчук М.Я.;1;Арсен Дедич;ООО Фирма Мелодия
ЖЗЛ;Митяев О.Г.;Митяев О.Г.;1;Олег Митяев;RMG Records
Я - Огонь, Ты - Вода;Брейтбург К.А.;Кавалерян К.А.;1;Марина Девятова;United Music Group
Три танкиста;Покрасс Д.Я.;Ласкин Б.С.;1;Виктор Селиванов;ООО Фирма Мелодия
Огонек;Народная;Исаковский М.В.;1;Владимир Нечаев;Апрелевский Завод
Беловежская пуща;Пахмутова А.Н.;Добронравов Н.Н.;1;Песняры;ООО Фирма Мелодия
Он не вернулся из боя;Высоцкий В.С.;Высоцкий В.С.;1;Владимир Высоцкий;ООО Фирма Мелодия
Время река;Айвазов А.Э.;Жуков В.П.;1;Александр Айвазов;ЗАО Квадро-Диск
5 минут до дома твоего;Кузьмин В.Б.;Кузьмин В.Б.;1;Владимир Кузьмин;Grand Records
Синий платочек;Петерсбурский Ю.Я.;Максимов М.А.;1;Клавдия Шульженко;Апрелевский Завод
Прадедушка;Ермолов А.В.;Загота М.А.;1;Непоседы;Immoral Basement Records
Моя любимая;Блантер М.И.;Долматовский Е.А.;1;Сергей Лемешев;Апрелевский Завод
Маки;Антонов Ю.М.;Поженяни Г.М.;1;Юрий Антонов;ООО Фирма Мелодия
Снегири;Антонов Ю.М.;Дудин М.А.;1;Юрий Антонов;ООО Фирма Мелодия
Небо;Макаревич А.Л.;Морсин А.Ю.;1;Лицей;Artur Music
Лучший парень;Потехин А.Е., Жуков С.Е.;Потехин А.Е., Жуков С.Е.;1;Руки вверх;Media Land
Минуты тишины;Петров А.П.;Матусовский М.Л.;1;Николай Караченцев;АРК-Систем Рекордз
За горизонт;Матвиенко И.И.;Матвиенко И.И., Жагун-Линник П.Н.;1;Фабрика & Иванушки Int;Монолит
Ветка жасмина;Дьячков С.К.;Ошанин Л.И.;1;Цветы;SNC Records
С добрым утром, любимая;Тарасов К.Б.;Митяев О.Г.;1;Олег Митяев;Star Mark
Изучай меня;Мелентьева Т.Г. (Сай);Мелентьева Т.Г. (Сай);1;Наталья Ветлицкая;Panorama Records
Вечер;Добрынин В. Г. (Антонов);Хайт А.И.;1;Лейся, песня;RDM
Пароход;Агутин Л.Н.;Агутин Л.Н.;1;Леонид Агутин;Iceberg Music
Жизнь;Антонов Ю.М.;Жуков О.В.;1;Юрий Антонов;ООО Фирма Мелодия
Уличные фонари;Татлян Ж.А.;Гарин Ю.М.(Герштейн);1;Валерий Сюткин;Мирумир
Час до рассвета;Дубравин Я.И.;Тальков И.В.;1;Людмила Сенчина, Группа Игоря Талькова;ООО Фирма Мелодия
Травы, травы;Шаинский В.Я.;Юшин И.С.;1;Геннадий Белов;ООО Фирма Мелодия
Ребята с нашего двора;Матвиенко И.И.;Шаганов А.А.;1;Любэ;Продюсерский Центр Игоря Матвиенко
25-й Этаж;Матвиенко И.И.;Жагун-Линник П.Н.;1;Корни;Mainstream Production
После дождя;Рябинин М.И.(Меерович);Фельцман О.Б.;1;Цветы;SNC Records
Чашка чаю;Добрынин В.Г.;Меерович М.И.;1;Весёлые ребята;Bomba Music
Сказка в моей жизни;Кузьмин В.Б.;Кузьмин В.Б.;1;Владимир Кузьмин;Real Records
Ты не одна;Клявер Д.И.;Костюшкин С.М.;1;Чай Вдвоём;Sound Pro
Если любишь ты;Антонов Ю.М.;Дербенев Л.П.;1;Веселые Ребята;Мелодия
Этот день;Газманов О.М.;Газманов О.М.;1;Олег Газманов;Rec Records
Березы;Матвиенко И.И.;Андреев М.В.;1;Любэ(Расторгуев Николай Вячеславович);Продюсерский центр Игоря Матвиенко
Районы кварталы;Билык Р.В.;Билык Р.В.;1;Звери;CD Land Records
Молния;Ковальский Д.В.;Ковальский Д.В.;1;Дима Билан;Archer Music Production
Прекрасное Далеко;Крылатов Е.П.;Энтин Ю.С.;1;Трио Меридиан;Feelee
Живет повсюду красота;Антонов Ю.М.;Пляцковский М.С.;1;Юрий Антонов;Star Track
Начни сначала;Мартынов Е.Г.;Вознесенский А.А.;1;Евгений Мартынов;ООО Фирма Мелодия"""

    data = [line.split(";") for line in raw_data.strip().split('\n')]
    out_path = "/home/blxnk/Documents/work/m.xlsx"

    generate_report(5, 2025, data, out_path)
