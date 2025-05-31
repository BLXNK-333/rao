from typing import List
import calendar
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _get_russian_period_text(month: int, year: int) -> str:
    month_names = {
        1: "января", 2: "февраля", 3: "марта", 4: "апреля",
        5: "мая", 6: "июня", 7: "июля", 8: "августа",
        9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
    }

    from_day = 1
    to_day = calendar.monthrange(year, month)[1]
    month_name = month_names[month]

    return f"{from_day} {month_name} {year} по {to_day} {month_name} {year}"


def generate_xlsx_month_report(
        month: int,
        year: int,
        data: List[List[str]],
        save_path: str | Path
):
    save_path = Path(save_path)
    # Инициализация книги и листа
    wb = Workbook()
    ws = wb.active
    ws.title = save_path.stem

    # Шрифты
    header_font = Font(name='Times New Roman', size=14, bold=True)
    subheader_font = Font(name='Times New Roman', size=12, bold=True)
    body_font = Font(name='Times New Roman', size=12)

    # Границы
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Ширина колонок
    col_widths = [48.71, 47.14, 44.14, 6.71, 44.0, 65.71]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Дата диапазон
    period_text = _get_russian_period_text(month, year)

    # Шапка
    header_text = (
        f"Отчет  об использованных фонограммах за период {period_text} в эфире радиоканала "
        "« Радио России Астрахань » - город федерального, регионального, республиканского "
        "либо краевого значения осуществляющего трансляцию"
    )
    ws.merge_cells("A1:F1")
    ws["A1"] = header_text
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 54  # увеличено, чтобы вместить многострочный текст

    # Заголовки таблицы
    table_headers = [
        "Название фонограммы",
        "Автор музыки",
        "Автор слов",
        "Кол - во сообщений в эфир",
        "Исполнитель (ФИО исполнителя или названия коллектива)",
        "Изготовитель фонограммы"
    ]
    for col, title in enumerate(table_headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 61.5  # под заголовки с переносом

    # Данные
    for row_idx, row_data in enumerate(data, start=3):
        ws.row_dimensions[row_idx].height = 29
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

    # Пустая строка отступа перед подвалом
    empty_row = len(data) + 3
    ws.row_dimensions[empty_row].height = 22

    # Подвал
    footer_lines = [
        "* ВГТРК ГТРК «Культура» (правопреемник Государственного дома радиовещания и "
        "звукозаписи ГДРЗ)",
        "* Self - Release –  творческая продукция исполнителя, доступная для продаж "
        "или распространения"
    ]
    for i, line in enumerate(footer_lines, start=empty_row + 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = body_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.row_dimensions[i].height = 22

    # Отступ после подвала — 2 строки по 22 pt
    post_footer_row1 = empty_row + len(footer_lines) + 1
    post_footer_row2 = post_footer_row1 + 1
    ws.row_dimensions[post_footer_row1].height = 22
    ws.row_dimensions[post_footer_row2].height = 22

    # Подписи
    signature_row = post_footer_row2 + 1
    ws.row_dimensions[signature_row].height = 22

    ws.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row,
                   end_column=3)  # A:C
    ws.merge_cells(start_row=signature_row, start_column=4, end_row=signature_row,
                   end_column=6)  # D:F

    ws.cell(row=signature_row, column=1,
            value="М.П.    _________________________   ").font = Font(
        name='Times New Roman', size=14, bold=True)
    ws.cell(row=signature_row, column=1).alignment = Alignment(horizontal="center",
                                                               vertical="bottom")

    ws.cell(row=signature_row, column=4, value="Директор ГТРК" + " " * 42).font = Font(
        name='Times New Roman', size=14, bold=True, underline="single")
    ws.cell(row=signature_row, column=4).alignment = Alignment(horizontal="center",
                                                               vertical="bottom")

    # Подписи: поясняющий текст
    explain_row = signature_row + 1
    ws.merge_cells(start_row=explain_row, start_column=1, end_row=explain_row,
                   end_column=3)  # A:C
    ws.merge_cells(start_row=explain_row, start_column=4, end_row=explain_row,
                   end_column=6)  # D:F

    ws.cell(row=explain_row, column=1, value="(подпись)").font = Font(
        name='Times New Roman', size=8)
    ws.cell(row=explain_row, column=1).alignment = Alignment(horizontal="center",
                                                             vertical="top")

    ws.cell(row=explain_row, column=4,
            value="(должность,  ФИО руководителя)").font = Font(
        name='Times New Roman', size=8)
    ws.cell(row=explain_row, column=4).alignment = Alignment(horizontal="center",
                                                             vertical="top")

    # Сохранение
    wb.save(save_path)